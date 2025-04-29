import os
import requests
import weaviate
import numpy as np
from sentence_transformers import SentenceTransformer
from dotenv import load_dotenv
import csv
from openpyxl import Workbook, load_workbook

load_dotenv()

# Constants
OLLAMA_URL = "http://localhost:11434/api/generate"
OLLAMA_MODEL = os.getenv("OLLAMA_MODEL", "llama3")
DATASET_PATH = os.getenv("DATASET_PATH", "qa_dataset_1000.csv")
OUTPUT_XLSX = "qa_with_predictions.xlsx"

# Connect to Weaviate
client = weaviate.Client("http://localhost:8080")

# Load embedding model
embed_model = SentenceTransformer("all-MiniLM-L6-v2")

# Retrieve context from Weaviate
def search_weaviate(query, top_k=3):
    query_embedding = embed_model.encode([query])
    query_embedding = np.array(query_embedding).astype(np.float32)

    response = client.query.get("WorldWarChunk", ["text"])\
        .with_near_vector({"vector": query_embedding[0]})\
        .with_limit(top_k).do()

    return [item["text"] for item in response["data"]["Get"]["WorldWarChunk"]]

# Query local LLM via Ollama
def query_llama(prompt):
    payload = {
        "model": OLLAMA_MODEL,
        "prompt": prompt,
        "stream": False
    }
    try:
        response = requests.post(OLLAMA_URL, json=payload)
        response.raise_for_status()
        return response.json().get("response", "").strip()
    except Exception as e:
        return f"Error querying {OLLAMA_MODEL}: {e}"

# Generator to fetch questions and answers
def question_fetch(csv_file_path):
    with open(csv_file_path, newline='', encoding='utf-8') as csvfile:
        reader = csv.reader(csvfile)
        header = next(reader, None)
        for row in reader:
            if row and len(row) >= 2:
                yield {"question": row[0], "answer": row[1]}

# Main
def main():
    print(f"\n🤖 World War History Q&A (Powered by {OLLAMA_MODEL})")

    # Create Excel workbook and sheet
    if os.path.exists(OUTPUT_XLSX):
        wb = load_workbook(OUTPUT_XLSX)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(["Question", "Answer", "predicted_answer", "context"])

    for item in question_fetch(DATASET_PATH):
        question = item["question"]
        ground_truth = item["answer"]

        if not question.strip():
            continue

        print(f"\n❓ Question: {question}")

        context_chunks = search_weaviate(question)
        context = "\n".join(context_chunks)[:1500]

        prompt = f"""
You are a factual Q&A assistant based on historical documents.

Instructions:
- Answer questions accurately and concisely using only the provided context.
- If the answer is a date, name, or simple fact, respond with just that, avoiding full sentences unless necessary.
- If the question contains a wrong assumption, correct it politely and provide the correct answer.
- If the answer is not found in the text, respond with: "Answer not found in the text."
- Do not make up information beyond what is given.

Context:
{context}

Question: {question}
Answer:
"""

        predicted_answer = query_llama(prompt)
        print(f"\n💬 Answer: {predicted_answer}\n")

        # Append to Excel
        ws.append([question, ground_truth, predicted_answer, context])
        wb.save(OUTPUT_XLSX)

    print(f"\n✅ All predictions saved to: {OUTPUT_XLSX}")

if __name__ == "__main__":
    main()
