import os
import requests
import weaviate
import numpy as np
import csv
import logging
from openpyxl import Workbook, load_workbook
from sentence_transformers import SentenceTransformer
from typing import Generator, Dict, List, Optional, Any

# Set up logging
current_dir = os.path.dirname(os.path.abspath(__file__))
output_dir = os.path.join(current_dir, 'output')
os.makedirs(output_dir, exist_ok=True)
log_file = os.path.join(output_dir, 'testing_data.log')

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    handlers=[
        logging.FileHandler(log_file),
        logging.StreamHandler()
    ]
)
logger = logging.getLogger(__name__)

# Get the correct paths using the project structure
project_root = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

# Constants
OLLAMA_URL = "http://localhost:11434/api/generate"
OLLAMA_MODEL = "llama3"
DATASET_PATH = os.path.join(project_root, "data", "qa_pairs", "qa_dataset_1000_part2.csv")
OUTPUT_XLSX = os.path.join(project_root, "data", "processed", "qa_with_predictions_part2.xlsx")

class RAGQuestionAnswering:
    """
    A class to perform retrieval-augmented generation (RAG) on a QA dataset,
    fetching context from Weaviate and generating answers using a local LLaMA model.
    
    Attributes:
        client (weaviate.Client): Client connection to Weaviate vector database.
        embed_model (SentenceTransformer): Model for creating text embeddings.
    """

    def __init__(self) -> None:
        """
        Initialize the RAG system with Weaviate client and embedding model.
        
        Raises:
            Exception: If connection to Weaviate fails or embedding model cannot be loaded.
        """
        try:
            self.client = weaviate.Client("http://localhost:8080")
        except Exception as e:
            logger.error(f"Error connecting to Weaviate: {e}")
            raise

        try:
            self.embed_model = SentenceTransformer("intfloat/e5-base-v2")
        except Exception as e:
            logger.error(f"Error loading embedding model: {e}")
            raise

    def search_weaviate(self, query: str, top_k: int = 3) -> List[str]:
        """
        Retrieves top_k relevant text chunks from Weaviate for a given query.
        
        Args:
            query (str): The question or search term to find relevant context for.
            top_k (int, optional): Number of text chunks to retrieve. Defaults to 3.
            
        Returns:
            List[str]: List of retrieved text chunks most relevant to the query.
            
        Note:
            Returns empty list if retrieval fails.
        """
        try:
            query_embedding = self.embed_model.encode([query])
            query_embedding = np.array(query_embedding).astype(np.float32)

            response = self.client.query.get("WorldWarChunk", ["text"])\
                .with_near_vector({"vector": query_embedding[0]})\
                .with_limit(top_k).do()

            return [item["text"] for item in response["data"]["Get"]["WorldWarChunk"]]
        except Exception as e:
            logger.error(f"Error querying Weaviate: {e}")
            return []

    def query_llama(self, prompt: str) -> str:
        """
        Sends a prompt to the local LLaMA model via Ollama API and returns the response.
        
        Args:
            prompt (str): The formatted prompt to send to the LLM.
            
        Returns:
            str: The generated response from LLaMA, or error message if query fails.
        """
        payload = {
            "model": OLLAMA_MODEL,
            "prompt": prompt,
            "stream": False
        }
        try:
            response = requests.post(OLLAMA_URL, json=payload)
            response.raise_for_status()
            return response.json().get("response", "").strip()
        except requests.exceptions.RequestException as e:
            logger.error(f"Error querying {OLLAMA_MODEL}: {e}")
            return f"Error querying {OLLAMA_MODEL}: {e}"

    def question_fetch(self, csv_file_path: str) -> Generator[Dict[str, str], None, None]:
        """
        Yields question-answer pairs from a CSV file with two columns: question and answer.
        
        Args:
            csv_file_path (str): Path to the CSV file containing QA pairs.
            
        Yields:
            Dict[str, str]: Dictionary with 'question' and 'answer' keys.
            
        Raises:
            Exception: If CSV file cannot be read or processed.
        """
        try:
            with open(csv_file_path, newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                next(reader, None)  # Skip header
                for row in reader:
                    if row and len(row) >= 2:
                        yield {"question": row[0], "answer": row[1]}
        except Exception as e:
            logger.error(f"Error reading CSV file {csv_file_path}: {e}")
            raise

    def run(self) -> None:
        """
        Executes the full RAG pipeline: fetch context, generate answer, write to Excel.
        
        Process:
            1. Opens or creates Excel output file
            2. Reads QA pairs from CSV file
            3. For each question, retrieves relevant context from Weaviate
            4. Generates answer using LLaMA model with the retrieved context
            5. Writes results to Excel file
            
        Raises:
            Exception: If file operations fail or other processing errors occur.
        """
        logger.info(f"🤖 World War History Q&A (Powered by {OLLAMA_MODEL})")

        if os.path.exists(OUTPUT_XLSX):
            try:
                wb = load_workbook(OUTPUT_XLSX)
                ws = wb.active
            except Exception as e:
                logger.error(f"Error opening Excel file {OUTPUT_XLSX}: {e}")
                raise
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Question", "Answer", "predicted_answer", "context"])

        for item in self.question_fetch(DATASET_PATH):
            question = item["question"]
            ground_truth = item["answer"]

            if not question.strip():
                continue

            logger.info(f"❓ Question: {question}")

            context_chunks = self.search_weaviate(question)
            context = "\n".join(context_chunks)[:1500]

            prompt = f"""
You are a factual Q&A assistant based on historical documents.

Instructions:
Give answers accurately and concisely based on the provided context.
Answers should be short and to the point.
Do not include any disclaimers or unnecessary information.
Do not make up any information.
Do not include any references to the source of the information.
Do not include any additional context or information outside of the provided context.

Context:
{context}

Question: {question}
Answer:
"""

            predicted_answer = self.query_llama(prompt)
            logger.info(f"💬 Answer: {predicted_answer}")

            ws.append([question, ground_truth, predicted_answer, context])
            try:
                wb.save(OUTPUT_XLSX)
            except Exception as e:
                logger.error(f"Error saving Excel file {OUTPUT_XLSX}: {e}")
                raise

        logger.info(f"✅ All predictions saved to: {OUTPUT_XLSX}")


if __name__ == "__main__":
    try:
        rag_qa = RAGQuestionAnswering()
        rag_qa.run()
    except Exception as e:
        logger.error(f"Error during script execution: {e}")
