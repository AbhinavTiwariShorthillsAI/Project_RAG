import os
import sys
import requests
import weaviate
import numpy as np
import csv
import logging
from openpyxl import Workbook, load_workbook
from sentence_transformers import SentenceTransformer
from typing import Generator, Dict, List, Optional, Any, Union

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

# Logging setup
logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")
logger = logging.getLogger(__name__)

# Constants
OLLAMA_URL = "http://localhost:11434/api/generate"
OLLAMA_MODEL = "llama3"
DATASET_PATH = "data/qa_dataset_1000_part2.csv"
OUTPUT_XLSX = "data/qa_with_predictions_part2.xlsx"

class RAGQuestionAnswering:
    """
    A class for Retrieval-Augmented Generation (RAG) on a QA dataset.
    
    This class connects to a Weaviate vector database, retrieves relevant contexts
    for questions, and generates answers using a local LLaMA model through Ollama.
    
    Attributes:
        client (weaviate.Client): Client connection to Weaviate database
        embed_model (SentenceTransformer): Model for generating text embeddings
    """

    def __init__(self) -> None:
        """
        Initialize the RAG system with Weaviate client and SentenceTransformer model.
        
        Raises:
            Exception: If connection to Weaviate fails or embedding model cannot be loaded
        """
        try:
            self.client = weaviate.Client("http://localhost:8080")
            logger.info("Connected to Weaviate successfully.")
        except Exception as e:
            logger.error(f"Error connecting to Weaviate: {e}")
            raise

        try:
            self.embed_model = SentenceTransformer("intfloat/e5-base-v2")
            logger.info("Embedding model loaded successfully.")
        except Exception as e:
            logger.error(f"Error loading embedding model: {e}")
            raise

    def search_weaviate(self, query: str, top_k: int = 3) -> List[str]:
        """
        Search Weaviate for text chunks relevant to the query.
        
        Args:
            query (str): The question or search query
            top_k (int, optional): Number of chunks to retrieve. Defaults to 3.
            
        Returns:
            List[str]: List of retrieved text chunks, or empty list if retrieval fails
        """
        try:
            query_embedding = self.embed_model.encode([query])
            query_embedding = np.array(query_embedding).astype(np.float32)

            response = self.client.query.get("WorldWarChunk", ["text"])\
                .with_near_vector({"vector": query_embedding[0]})\
                .with_limit(top_k).do()

            results = [item["text"] for item in response["data"]["Get"]["WorldWarChunk"]]
            logger.info(f"Retrieved {len(results)} chunks from Weaviate for query: {query}")
            return results
        except Exception as e:
            logger.error(f"Error querying Weaviate: {e}")
            return []

    def query_llama(self, prompt: str) -> str:
        """
        Query the LLaMA model via Ollama API with a prompt.
        
        Args:
            prompt (str): The formatted prompt to send to the model
            
        Returns:
            str: The model's response or an error message if the query fails
        """
        payload = {
            "model": OLLAMA_MODEL,
            "prompt": prompt,
            "stream": False
        }
        try:
            response = requests.post(OLLAMA_URL, json=payload)
            response.raise_for_status()
            answer = response.json().get("response", "").strip()
            logger.info("Received response from LLaMA model.")
            return answer
        except requests.exceptions.RequestException as e:
            logger.error(f"Error querying {OLLAMA_MODEL}: {e}")
            return f"Error querying {OLLAMA_MODEL}: {e}"

    def question_fetch(self, csv_file_path: str) -> Generator[Dict[str, str], None, None]:
        """
        Fetch question-answer pairs from a CSV file.
        
        Args:
            csv_file_path (str): Path to the CSV file containing QA pairs
            
        Yields:
            Dict[str, str]: Dictionary with 'question' and 'answer' keys
            
        Raises:
            Exception: If the CSV file cannot be read or processed
        """
        try:
            with open(csv_file_path, newline='', encoding='utf-8') as csvfile:
                reader = csv.reader(csvfile)
                next(reader, None)
                for row in reader:
                    if row and len(row) >= 2:
                        yield {"question": row[0], "answer": row[1]}
        except Exception as e:
            logger.error(f"Error reading CSV file {csv_file_path}: {e}")
            raise

    def run(self) -> None:
        """
        Execute the complete RAG pipeline.
        
        Process:
        1. Sets up output Excel file
        2. Reads questions from the dataset
        3. For each question, retrieves relevant context
        4. Generates an answer using the LLaMA model
        5. Saves results to the Excel file
        
        Raises:
            Exception: If file operations fail or other errors occur during processing
        """
        logger.info("Starting RAG Question Answering process.")
        logger.info(f"Using dataset: {DATASET_PATH}")

        if os.path.exists(OUTPUT_XLSX):
            try:
                wb = load_workbook(OUTPUT_XLSX)
                ws = wb.active
                logger.info(f"Loaded existing Excel file: {OUTPUT_XLSX}")
            except Exception as e:
                logger.error(f"Error opening Excel file {OUTPUT_XLSX}: {e}")
                raise
        else:
            wb = Workbook()
            ws = wb.active
            ws.append(["Question", "Answer", "predicted_answer", "context"])
            logger.info(f"Created new Excel file: {OUTPUT_XLSX}")

        for item in self.question_fetch(DATASET_PATH):
            question = item["question"]
            ground_truth = item["answer"]

            if not question.strip():
                logger.warning("Encountered empty question. Skipping.")
                continue

            logger.info(f"Processing question: {question}")

            context_chunks = self.search_weaviate(question)
            context = "\n".join(context_chunks)[:1500]

            prompt = f"""
You are a factual Q&A assistant based on historical documents.

Instructions:
Give answers accurately and concisely based on the provided context.
Answers should be short and to the point.
Do not include any disclaimers or unnecessary information.
Do not make up any information.
Do not include any references to the source of the information.
Do not include any additional context or information outside of the provided context.

Context:
{context}

Question: {question}
Answer:
"""

            predicted_answer = self.query_llama(prompt)
            logger.info(f"Predicted answer: {predicted_answer[:100]}...")

            ws.append([question, ground_truth, predicted_answer, context])
            try:
                wb.save(OUTPUT_XLSX)
            except Exception as e:
                logger.error(f"Error saving Excel file {OUTPUT_XLSX}: {e}")
                raise

        logger.info(f"All predictions saved to: {OUTPUT_XLSX}")
        logger.info("RAG pipeline completed successfully.")


if __name__ == "__main__":
    try:
        rag_qa = RAGQuestionAnswering()
        rag_qa.run()
    except Exception as e:
        logger.error(f"Fatal error during script execution: {e}")
